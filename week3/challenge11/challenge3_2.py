# -*- coding: utf-8 -*-

from openpyxl import load_workbook,Workbook


def combine():
    wb = load_workbook('courses.xlsx')
    ws1 = wb.get_sheet_by_name('students')
    ws2 = wb.get_sheet_by_name('time')
    dict1 = {}
    dict2 = {}
    for i in range(2,ws1.max_row+1):
        dict1.setdefault(ws1.cell(row=i,column=2).value,[0,0])
        dict1[ws1.cell(row=i,column=2).value] = [ws1.cell(row=i,column=1).value, ws1.cell(row=i,column=3).value]

    for i in range(2,ws2.max_row+1):
        dict2.setdefault(ws2.cell(row=i,column=2).value,[0,0])
        dict2[ws2.cell(row=i,column=2).value] = [ws2.cell(row=i,column=1).value, ws2.cell(row=i,column=3).value]
#    print(dict1)
#    print(dict2)
    ws3 = wb.create_sheet(title='combine')
    ws3.append(['创建时间','课程名称','学习人数','学习时间'])
    for k, v in dict1.items():
        for m, n in dict2.items():
            if k == m:
                ws3.append([v[0],k,v[1],n[1]])

    wb.save('courses.xlsx')


def split():
    wb = load_workbook('courses.xlsx')
    ws = wb.get_sheet_by_name('combine')
    s = set()
    l = list(ws.values)[1:]
    for i in l:
        s.add(i[0].strftime('%Y'))
    for y in s:
        wb2 = Workbook()
        ws4 = wb2.active
        ws4.title = y
        for i in l:
            if i[0].strftime('%Y') == y:
                ws4.append(i)
        wb2.save('{}.xlsx'.format(y))

    
if __name__ == '__main__':
    combine()
    split()
